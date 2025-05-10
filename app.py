from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort, jsonify, session, g
from flask_login import current_user
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime
from pymongo import MongoClient
from bson import ObjectId
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pymongo import TEXT
from math import ceil
from jinja2 import Environment
from functools import wraps

load_dotenv()

def setup_jinja_filters(app):
    @app.template_filter('zip_lists')
    def zip_lists(a, b):
        return zip(a, b)
        
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'f5b0c2b44c0da7ddca709558e7fa4e9f9fb59d65144d7683')

client = MongoClient(os.getenv('MONGODB_URI', 'mongodb+srv://Admin:060843Za@summary.ezehedg.mongodb.net/'))
db = client['finance_web_app']

app.jinja_env.globals.update(zip=zip)

# ข้อมูลบัญชีธนาคาร
BANK_ACCOUNTS = {
    'BBL': 'ธนาคารกรุงเทพ',
    'KBANK': 'ธนาคารกสิกรไทย',
    'KBIZ': 'ธนาคารกสิกรไทย',
    'KTB': 'ธนาคารกรุงไทย',
    'SCB': 'ธนาคารไทยพาณิชย์',
    'BAY': 'ธนาคารกรุงศรีอยุธยา',
    'TMB': 'ธนาคารทหารไทย',
    'TBANK': 'ธนาคารธนชาต',
    'KK': 'ธนาคารเกียรตินาคิน',
    'TISCO': 'ธนาคารทิสโก้',
    'CIMBT': 'ธนาคารซีไอเอ็มบีไทย',
    'LH': 'ธนาคารแลนด์ แอนด์ เฮ้าส์',
    'UOB': 'ธนาคารยูโอบี',
    'BACC': 'ธนาคารเพื่อการเกษตร',
    'ICBC': 'ธนาคารไอซีบีซี',
    'GSB': 'ธนาคารออมสิน',
    'TRUE': 'ทรูมันนี่วอเล็ต'
}

with app.app_context():
    db.transactions.create_index([('agent', TEXT)])

setup_jinja_filters(app)

@app.context_processor
def inject_datetime():
    return {'datetime': datetime, 'bank_accounts': BANK_ACCOUNTS}

@app.template_filter('number_format')
def number_format(value, precision=2):
    try:
        if value is None:
            return f"{0.0:,.{precision}f}"
        return f"{float(value):,.{precision}f}"
    except (ValueError, TypeError):
        return f"{0.0:,.{precision}f}"


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(role):
    def wrapper(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            user = db.users.find_one({'_id': ObjectId(session['user_id'])})
            if user['role'] != role:
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return wrapper

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # ค้นหาผู้ใช้ในฐานข้อมูล
        user = db.users.find_one({'username': username})
        
        # ตรวจสอบรหัสผ่าน
        if user and check_password_hash(user['password'], password):
            # บันทึกข้อมูลผู้ใช้ใน session
            session['user_id'] = str(user['_id'])
            session['username'] = user['username']
            flash('เข้าสู่ระบบสำเร็จ', 'success')
            return redirect(url_for('index'))  # เปลี่ยนเส้นทางไปหน้า index.html
        else:
            flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)  # ลบข้อมูล session
    flash('ออกจากระบบสำเร็จ', 'success')  # แสดงข้อความออกจากระบบ
    return redirect(url_for('login'))  # เปลี่ยนเส้นทางไปยังหน้า login

@app.route('/')
@login_required
def index():
    # ตรวจสอบว่า user_id อยู่ใน session หรือไม่
    if 'user_id' not in session:
        return redirect(url_for('login'))  # ถ้าไม่มีก็ไปที่หน้า login

    # ดึงข้อมูลผู้ใช้จากฐานข้อมูล
    user = db.users.find_one({'_id': ObjectId(session['user_id'])})

    if user:
        # ดึงข้อมูล agent ทั้งหมด
        agents = list(db.agents.find())
        return render_template('index.html', agents=agents, user=user)  # ส่งข้อมูลผู้ใช้และ agent ไปยังเทมเพลต
    else:
        # ถ้าไม่พบผู้ใช้ ให้ไปที่หน้า login
        return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required  # ทำให้เฉพาะผู้ใช้ที่ล็อกอินสามารถเข้าถึงได้
def change_password():
    if request.method == 'POST':
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        confirm_new_password = request.form['confirm_new_password']

        # ตรวจสอบว่ารหัสผ่านใหม่กับรหัสผ่านยืนยันใหม่ตรงกัน
        if new_password != confirm_new_password:
            flash('รหัสผ่านใหม่ไม่ตรงกัน', 'danger')
            return redirect(url_for('change_password'))

        # ตรวจสอบว่ารหัสผ่านเก่าถูกต้องหรือไม่
        user = db.users.find_one({'_id': ObjectId(session['user_id'])})
        if not user or not check_password_hash(user['password'], old_password):
            flash('รหัสผ่านเก่าไม่ถูกต้อง', 'danger')
            return redirect(url_for('change_password'))

        # อัปเดตรหัสผ่านใหม่ในฐานข้อมูล
        hashed_password = generate_password_hash(new_password)
        db.users.update_one({'_id': ObjectId(session['user_id'])}, {'$set': {'password': hashed_password}})
        flash('เปลี่ยนรหัสผ่านสำเร็จ', 'success')
        return redirect(url_for('index'))  # เปลี่ยนเส้นทางกลับไปหน้า index หรือหน้าอื่นๆ ที่ต้องการ

    return render_template('change_password.html')  # แสดงฟอร์มการเปลี่ยนรหัสผ่าน

@app.route('/admin-panel', methods=['GET', 'POST'])
@login_required
@role_required('admin')  # ตรวจสอบว่าผู้ใช้เป็น admin หรือไม่
def admin_panel():
    # ดึงข้อมูลผู้ใช้ทั้งหมดจากฐานข้อมูล
    users = db.users.find()

    if request.method == 'POST':
        action = request.form.get('action')

        # ฟังก์ชันในการเพิ่มผู้ใช้ใหม่
        if action == 'add':
            username = request.form['username']
            password = request.form['password']
            role = request.form['role']

            if username and password and role:
                # สร้างรหัสผ่านที่เข้ารหัส
                hashed_password = generate_password_hash(password)
                
                # เพิ่มผู้ใช้ใหม่ในฐานข้อมูล
                db.users.insert_one({
                    'username': username,
                    'password': hashed_password,
                    'role': role,  # กำหนดบทบาท (role)
                    'created_at': datetime.now()
                })
                flash('เพิ่มผู้ใช้ใหม่สำเร็จ', 'success')
            else:
                flash('กรุณากรอกข้อมูลให้ครบถ้วน', 'danger')

        # ฟังก์ชันในการแก้ไขข้อมูลผู้ใช้
        if action == 'edit':
            user_id = request.form['user_id']
            username = request.form['username']
            role = request.form['role']
            new_password = request.form.get('new_password')

            # ตรวจสอบว่ามีการกรอกรหัสผ่านใหม่หรือไม่
            update_data = {'username': username, 'role': role}

            if new_password:
                # ถ้ามีการกรอกรหัสผ่านใหม่ จะทำการเข้ารหัสและอัปเดต
                hashed_password = generate_password_hash(new_password)
                update_data['password'] = hashed_password

            # อัปเดตข้อมูลผู้ใช้
            db.users.update_one(
                {'_id': ObjectId(user_id)},
                {'$set': update_data}
            )
            flash('แก้ไขข้อมูลผู้ใช้สำเร็จ', 'success')
        
        return redirect(url_for('admin_panel'))

    return render_template('admin_panel.html', users=users)

@app.route('/submit', methods=['POST'])
@login_required
def submit_transaction():
    try:
        # รับข้อมูลบัญชีแบบไดนามิก
        deposit_accounts = request.form.getlist('deposit_accounts[]')
        deposit_amounts = [float(x) for x in request.form.getlist('deposit_amounts[]')]
        withdrawal_accounts = request.form.getlist('withdrawal_accounts[]')
        withdrawal_amounts = [float(x) for x in request.form.getlist('withdrawal_amounts[]')]

        transaction_data = {
            'agent': request.form.get('agent'),
            'date': datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
            'new_members': int(request.form.get('new_members', 0)),
            'first_deposit': float(request.form.get('first_deposit', 0)),
            'all_deposit': float(request.form.get('all_deposit', 0)),
            'used': int(request.form.get('used', 0)),
            'db': float(request.form.get('db', 0)),
            'deposit_accounts': deposit_accounts,
            'deposit_amounts': deposit_amounts,
            'withdrawal_accounts': withdrawal_accounts,
            'withdrawal_amounts': withdrawal_amounts,
            'profit_thb': float(request.form.get('profit_thb', 0)),
            'profit_usdt': float(request.form.get('profit_usdt', 0)),
            'transfer': float(request.form.get('transfer', 0)),
            'mobile_deposit': float(request.form.get('mobile_deposit', 0)),
            'bonus': float(request.form.get('bonus', 0)),
            'commission': float(request.form.get('commission', 0)),
            'admin_approved': int(request.form.get('admin_approved', 0)),
            'created_at': datetime.now(),
            'created_by': session.get('user_id')
        }

        if not transaction_data['agent']:
            flash('กรุณาเลือก Agent', 'danger')
            return redirect(url_for('index'))

        db.transactions.insert_one(transaction_data)
        flash('บันทึกข้อมูลสำเร็จ', 'success')
        return redirect(url_for('show_history'))

    except Exception as e:
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/history')
def show_history():
    query = {}
    page = request.args.get('page', 1, type=int)
    per_page = 31

    selected_agent = request.args.get('agent', '')
    selected_month = request.args.get('month', '')
    selected_year = request.args.get('year', '')

    if selected_agent:
        query['agent'] = selected_agent
    
    if selected_month and selected_year:
        try:
            start_date = datetime(int(selected_year), int(selected_month), 1)
            end_month = int(selected_month) + 1 if int(selected_month) != 12 else 1
            end_year = int(selected_year) if int(selected_month) != 12 else int(selected_year) + 1
            end_date = datetime(end_year, end_month, 1)
            query['date'] = {'$gte': start_date, '$lt': end_date}
        except ValueError:
            flash('รูปแบบเดือน/ปีไม่ถูกต้อง', 'danger')

    unique_agents = db.transactions.distinct('agent')
    transactions_cursor = db.transactions.find(query).sort('date', -1)
    total = db.transactions.count_documents(query)
    transactions_list = list(transactions_cursor.skip((page - 1) * per_page).limit(per_page))

    # ตรวจสอบและตั้งค่าเริ่มต้นสำหรับ transaction แต่ละรายการ
    for transaction in transactions_list:
        transaction.setdefault('mobile_deposit', 0)
        transaction.setdefault('commission', 0)
        transaction.setdefault('bonus', 0)
        transaction.setdefault('transfer', 0)
        transaction.setdefault('db', 0)
        transaction.setdefault('deposit_accounts', [])
        transaction.setdefault('deposit_amounts', [])
        transaction.setdefault('withdrawal_accounts', [])
        transaction.setdefault('withdrawal_amounts', [])

        # ดึงชื่อผู้ใช้ที่รับผิดชอบจาก session
        user_id = transaction.get('created_by')
        if user_id:
            user = db.users.find_one({'_id': ObjectId(user_id)})
            transaction['created_by_name'] = user.get('username', 'Unknown')

    total_deposit = sum(sum(t.get('deposit_amounts', [])) for t in transactions_list)
    total_withdrawal = sum(sum(t.get('withdrawal_amounts', [])) for t in transactions_list)
    total_commission = sum(t.get('commission', 0) for t in transactions_list)
    total_bonus = sum(t.get('bonus', 0) for t in transactions_list)
    balance = total_deposit - (total_withdrawal + total_commission + total_bonus)

    months_years = list(db.transactions.aggregate([
        {"$group": {
            "_id": {
                "year": {"$year": "$date"},
                "month": {"$month": "$date"}
            }
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]))

    class Pagination:
        def __init__(self, page, per_page, total, items):
            self.page = page
            self.per_page = per_page
            self.total = total
            self.items = items
            self.pages = (total // per_page) + (1 if total % per_page > 0 else 0)
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

        def iter_pages(self, left_edge=2, left_current=2, right_current=5, right_edge=2):
            last = 0
            for num in range(1, self.pages + 1):
                if (num <= left_edge or 
                    (num > self.page - left_current - 1 and num < self.page + right_current) or 
                    num > self.pages - right_edge):
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num

    pagination = Pagination(
        page=page,
        per_page=per_page,
        total=total,
        items=transactions_list
    )
    
    return render_template('history.html',
                        unique_agents=unique_agents,
                        selected_agent=selected_agent,
                        transactions=pagination,
                        total_deposit=total_deposit,
                        total_withdrawal=total_withdrawal,
                        total_commission=total_commission,
                        total_bonus=total_bonus,
                        balance=balance,
                        months_years=months_years,
                        selected_month=selected_month,
                        selected_year=selected_year)

@app.before_request
def before_request():
    if 'user_id' in session:
        # ดึงข้อมูลผู้ใช้จากฐานข้อมูล
        user = db.users.find_one({'_id': ObjectId(session['user_id'])})
        g.user = user  # เก็บข้อมูลผู้ใช้ใน g เพื่อให้เข้าถึงได้ในทุกๆ request
    else:
        g.user = None

@app.route('/agents', methods=['GET', 'POST'])
@login_required
@role_required('admin')  # ตรวจสอบบทบาทของผู้ใช้เป็น admin
def manage_agents():
    if request.method == 'POST':
        action = request.form.get('action')

        # การเพิ่ม Agent
        if action == 'add':
            agent_name = request.form.get('name')
            if agent_name:
                db.agents.insert_one({
                    'name': agent_name,
                    'created_at': datetime.now()
                })
                flash('เพิ่ม Agent สำเร็จ', 'success')
            else:
                flash('กรุณากรอกชื่อ Agent', 'danger')

        # การลบ Agent
        elif action == 'delete':
            try:
                agent_id = request.form.get('agent_id')

                # ตรวจสอบว่า agent_id เป็น ObjectId หรือไม่
                if ObjectId.is_valid(agent_id):
                    agent_id = ObjectId(agent_id)
                    result = db.agents.delete_one({'_id': agent_id})
                    if result.deleted_count > 0:
                        flash('ลบ Agent สำเร็จ', 'success')
                    else:
                        flash('ไม่พบ Agent ที่ต้องการลบ', 'danger')
                else:
                    flash('ID ของ Agent ไม่ถูกต้อง', 'danger')

            except Exception as e:
                flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

        return redirect(url_for('manage_agents'))

    # ดึงข้อมูล Agent ทั้งหมด
    agents = list(db.agents.find().sort('name', 1))
    return render_template('agents.html', agents=agents)

@app.route('/edit-transaction/<transaction_id>', methods=['GET', 'POST'])
def edit_transaction(transaction_id):
    try:
        obj_id = ObjectId(transaction_id)
        transaction = db.transactions.find_one({'_id': obj_id})
        
        if not transaction:
            flash('ไม่พบรายการนี้ในระบบ', 'danger')
            return redirect(url_for('show_history'))

        if request.method == 'POST':
            # รับข้อมูลบัญชีแบบไดนามิก
            deposit_accounts = request.form.getlist('deposit_accounts[]')
            deposit_amounts = [float(x) for x in request.form.getlist('deposit_amounts[]')]
            withdrawal_accounts = request.form.getlist('withdrawal_accounts[]')
            withdrawal_amounts = [float(x) for x in request.form.getlist('withdrawal_amounts[]')]

            update_data = {
                'agent': request.form.get('agent'),
                'date': datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
                'new_members': int(request.form.get('new_members', 0)),
                'first_deposit': float(request.form.get('first_deposit', 0)),
                'all_deposit': float(request.form.get('all_deposit', 0)),
                'used': int(request.form.get('used', 0)),
                'db': float(request.form.get('db', 0)),
                'deposit_accounts': deposit_accounts,
                'deposit_amounts': deposit_amounts,
                'withdrawal_accounts': withdrawal_accounts,
                'withdrawal_amounts': withdrawal_amounts,
                'profit_thb': float(request.form.get('profit_thb', 0)),
                'profit_usdt': float(request.form.get('profit_usdt', 0)),
                'transfer': float(request.form.get('transfer', 0)),
                'mobile_deposit': float(request.form.get('mobile_deposit', 0)),
                'bonus': float(request.form.get('bonus', 0)),
                'commission': float(request.form.get('commission', 0)),
                'admin_approved': int(request.form.get('admin_approved', 0)),
                'updated_at': datetime.now()
            }

            db.transactions.update_one(
                {'_id': obj_id},
                {'$set': update_data}
            )
            
            flash('อัปเดตข้อมูลสำเร็จ', 'success')
            return redirect(url_for('show_history'))

        agents = list(db.agents.find())
        return render_template('edit_transaction.html', 
                            transaction=transaction, 
                            agents=agents,
                            datetime=datetime)

    except Exception as e:
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')
        return redirect(url_for('show_history'))

@app.route('/delete-transaction/<transaction_id>', methods=['POST'])
@login_required
@role_required('admin')  # เพิ่มการตรวจสอบยศเป็นแอดมิน
def delete_transaction(transaction_id):
    try:
        # แปลง ID ของรายการให้เป็น ObjectId
        obj_id = ObjectId(transaction_id)
        
        # ลบรายการจากฐานข้อมูล
        result = db.transactions.delete_one({'_id': obj_id})
        
        # ตรวจสอบว่ามีการลบรายการหรือไม่
        if result.deleted_count > 0:
            flash('ลบรายการสำเร็จ', 'success')
        else:
            flash('ไม่พบรายการที่ต้องการลบ', 'danger')
    
    except Exception as e:
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')
    
    return redirect(url_for('show_history'))

@app.route('/export/<transaction_id>')
def export_excel(transaction_id):
    try:
        transaction = db.transactions.find_one({'_id': ObjectId(transaction_id)})
        
        if not transaction:
            flash('ไม่พบรายการนี้ในระบบ', 'danger')
            return redirect(url_for('show_history'))

        # สร้างไฟล์ Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "สรุปยอดปิดกะ"

        # สไตล์และรูปแบบ
        header_fill = PatternFill(start_color='FF9900', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='D9EAD3', fill_type='solid')
        red_fill = PatternFill(start_color='F4CCCC', fill_type='solid')
        gray_fill = PatternFill(start_color='F3F3F3', fill_type='solid')
        blue_fill = PatternFill(start_color='c9daf8', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # สไตล์ฟอนต์
        font_header = Font(name='Arial', size=10, bold=False)
        font_normal = Font(name='Arial', size=10)
        font_date = Font(name='Arial', size=8)
        thai_date = f"{transaction['date'].day}/{transaction['date'].month}/{transaction['date'].year + 543}"


        # ส่วนหัวหลัก
        ws.merge_cells('A1:N1')
        ws['A1'] = f"สรุปยอดปิดกะ {transaction['agent']} วันที่ {transaction['date'].strftime('%d/%m/%Y')}"
        ws['A1'].font = Font(name="K2D", bold=True, size=15)
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # ส่วนหัวตาราง
        headers = [
            "วัน", "ฝาก", "เติมมือ", "ถอน", "ถอนมือ", 
            "ขาด/เกิน", "กำไร", "สมัคร", "เติมจริง", 
            "ไม่เติม", "ฝากแรก", "แอคทีฟ", "โบนัส", "ค่าคอม"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = yellow_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font_header

        # คำนวณยอดรวมจากบัญชีแบบไดนามิก
        total_deposit = sum(transaction.get('deposit_amounts', []))
        total_withdrawal = sum(transaction.get('withdrawal_amounts', []))
        atm_balance = total_deposit - (total_withdrawal + transaction.get('commission', 0))
        shortfall = transaction.get('transfer', 0) - atm_balance

        data_row = [
            thai_date,
            total_deposit,
            transaction.get('mobile_deposit', 0),
            total_withdrawal,
            total_withdrawal,
            shortfall,
            total_deposit - (total_withdrawal + transaction.get('commission', 0)) + shortfall,
            transaction.get('new_members', 0),
            transaction.get('first_deposit', 0),
            transaction.get('new_members', 0) - transaction.get('first_deposit', 0),
            transaction.get('all_deposit', 0),
            transaction.get('used', 0),
            transaction.get('bonus', 0),
            transaction.get('commission', 0)
        ]

        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=3, column=col_num)
            
            # กำหนดสูตรและค่า
            if col_num == 5:    # E3
                cell.value = '=SUM(D3)'
                cell.number_format = '#,##0.00'
            elif col_num == 7: # G3
                cell.value = '=SUM(B3-(D3+N3)+F3)'
            elif col_num == 10: # J3
                cell.value = '=SUM(H3-I3)'
            else:
                cell.value = value
            
            # กำหนดสไตล์
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # กำหนดฟอนต์
            if col_num == 1:  # A3
                cell.font = font_date
            else:
                cell.font = font_normal

            # กำหนดรูปแบบตัวเลข
            if col_num in [2, 3, 4, 5, 6, 7, 11, 13]:  # B2 C3, D3, E3
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0'


            if col_num in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 12, 14]:  # B2 C3, D3, E3
                cell.alignment = Alignment(horizontal='center', vertical='center')

            
            # กำหนดสีพื้นหลัง
            if col_num in [2, 3, 5, 8, 9 , 10, 11, 12, 14]:  # C3, D3, E3
                cell.fill = gray_fill
            elif col_num == 4:    # คอลัมน์ D
                cell.fill = red_fill
            elif col_num == 13: # คอลัมน์ M
                cell.fill = blue_fill
            elif col_num in [6,7]: # คอลัมน์ F และ G
                cell.fill = green_fill if (value >= 0 if col_num != 5 else cell.value >= 0) else red_fill
            else:
                cell.fill = gray_fill

        # ปรับความกว้างคอลัมน์
        column_widths = [15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ส่งไฟล์กลับ
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"สรุปยอด_{transaction['agent']}_{transaction['date'].strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'เกิดข้อผิดพลาดในการส่งออกไฟล์: {str(e)}', 'danger')
        return redirect(url_for('show_history'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)